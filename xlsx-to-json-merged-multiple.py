from json import dumps, load
from os.path import exists
from os import mkdir

from common import checkFolders
from openpyxl import load_workbook
from math import floor
from pydash import merge
from datetime import datetime

timestamp = floor(datetime.timestamp(datetime.now()))

checkFolders()

xlsxPath = './inputs/translation.xlsx'


def outputPath(region):
    if (exists(f'./outputs/merged-translations-multiple/{timestamp}/{region}') == False):
        mkdir(f'./outputs/merged-translations-multiple/{timestamp}/{region}')
    return f'./outputs/merged-translations-multiple/{timestamp}/{region}/translation.json'


def inputPath(region):
    return f'./inputs-multiple/{region}/translation.json'


def convertInJson(trans):
    objecto = {f'{trans["language"]}': {'translation': {}}}
    for keyl in list(trans['trans'].keys()):
        counter = objecto[f'{trans["language"]}']['translation']
        pathForJson = keyl.split('.')
        for i, key in enumerate(pathForJson):
            if i < len(pathForJson) - 1:
                if (counter.get(key, None) != None):
                    counter[key] = {**counter[key]}
                else:
                    counter[key] = {}
            else:
                counter[key] = trans['trans'][keyl]
            counter = counter[key]
    return objecto


if exists(xlsxPath):
    print(f'[READING] {xlsxPath}')
    wb = load_workbook(xlsxPath)
    ws = wb.worksheets[0]
    translationIdsColumn = list(ws['A'])
    del translationIdsColumn[0]
    languages = list(ws[1])
    del languages[0]
    for index in range(len(languages)):
        if languages[index].value.lower() == 'note':
            del languages[index]
    translations = []
    for language in languages:
        coupled = {}
        for identifier in translationIdsColumn:
            coupled[identifier.value] = ws.cell(
                row=identifier.row, column=language.column).value
        trans = convertInJson(
            {'language': language.value, 'trans': coupled})
        translations.append({'language': language.value, 'trans': trans})

    if (exists(f'./outputs/merged-translations-multiple/{timestamp}') == False):
        mkdir(f'./outputs/merged-translations-multiple/{timestamp}')

    for translation in translations:
        if exists(inputPath(translation['language'])):
            with open(inputPath(translation['language']), 'r') as f:
                data = load(f)
                with open(outputPath(translation['language']), 'w') as outfile:
                    outfile.write(dumps(merge(data, translation['trans'])))
        else:
            with open(outputPath(translation['language']), 'w') as outfile:
                outfile.write(dumps(translation['trans']))

else:
    print(f'[ERROR] no file at {xlsxPath}')
