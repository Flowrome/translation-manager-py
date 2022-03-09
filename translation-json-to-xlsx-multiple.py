from datetime import datetime
from json import load
from os.path import exists

from common import checkFolders
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.cell.cell import Cell
from math import floor
from glob import glob

checkFolders()

jsonPath = './inputs-multiple/**/translation.json'
outputPath = f'./outputs/exported-xlsx-multiple/translation-{floor(datetime.timestamp(datetime.now()))}.xlsx'


def createWorkbook(j):
    wb = Workbook()
    ws = wb.active
    translationId = Cell(ws, value='Translation ID', column="A", row=1)
    translationId.border = Border(left=Side(border_style='thin',  color='FF000000'), right=Side(
        border_style='thin', color='FF000000'), top=Side(border_style='thin', color='FF000000'), bottom=Side(border_style='thin'))
    translationId.font = Font(b=True, color='ffffff')
    translationId.fill = PatternFill("solid", fgColor='000000')
    colums = [translationId]
    for item in j:
        cell = Cell(ws, value=item['region'], column="A", row=1)
        cell.font = Font(b=True, color='ffffff')
        cell.fill = PatternFill("solid", fgColor='000000')
        cell.border = Border(left=Side(border_style='thin',  color='FF000000'), right=Side(
            border_style='thin', color='FF000000'), top=Side(border_style='thin', color='FF000000'), bottom=Side(border_style='thin'))
        colums.append(cell)
    ws.append(colums)
    keys = []
    for k in j[0]['normalized']:
        keys.append(k)
    matrix = []
    for key in keys:
        keyCell = Cell(ws, value=key, column="A", row=1)
        keyCell.border = Border(left=Side(border_style='thin',  color='FF000000'), right=Side(
            border_style='thin', color='FF000000'), top=Side(border_style='thin', color='FF000000'), bottom=Side(border_style='thin'))
        counter = [keyCell]
        for item in j:
            normalized = item['normalized']
            cell = Cell(ws, value=normalized[key], column="A", row=1)
            cell.border = Border(left=Side(border_style='thin',  color='FF000000'), right=Side(
                border_style='thin', color='FF000000'), top=Side(border_style='thin', color='FF000000'), bottom=Side(border_style='thin'))
            counter.append(cell)
        matrix.append(counter)

    for elem in matrix:
        ws.append(elem)

    for col in ws.columns:
        column = col[0].column_letter
        ws.column_dimensions[column].width = 45
    wb.save(outputPath)


def flatten_json(y):
    out = {}

    def flatten(x, name=''):

        # If the Nested key-value
        # pair is of dict type
        if type(x) is dict:

            for a in x:
                flatten(x[a], name + a + '.')

        # If the Nested key-value
        # pair is of list type
        elif type(x) is list:

            i = 0

            for a in x:
                flatten(a, name + str(i) + '.')
                i += 1
        else:
            out[name[:-1]] = x

    flatten(y)
    return out


if glob(jsonPath, recursive=True) != []:
    globs = glob(jsonPath, recursive=True)
    data_regions = []
    for path in globs:
        print(f'[READING] {path}')
        with open(path, 'r') as f:
            data = load(f)
            if list(data[list(data.keys())[0]].keys())[0] == 'translation':
                region = list(data.keys())[0]
                normalized = flatten_json(data[list(data.keys())[0]]
                                          [list(data[list(data.keys())[0]].keys())[0]])
                final_object = {'region': region, 'normalized': normalized}
                data_regions.append(final_object)

    createWorkbook(data_regions)
else:
    print(f'[ERROR] no file at {jsonPath}')
