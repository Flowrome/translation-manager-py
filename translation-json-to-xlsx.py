from datetime import datetime
from json import load
from os.path import exists

from common import checkFolders
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from math import floor

checkFolders()

jsonPath = './inputs/translation.json'
outputPath = f'./outputs/exported-xlsx/translation-{floor(datetime.timestamp(datetime.now()))}.xlsx'


def createWorkbook(j):
    wb = Workbook()
    ws = wb.active
    ws.append(['Translation ID', 'Value'])
    translationIds = ws['A1']
    values = ws['B1']
    translationIds.font = Font(b=True, color='ffffff')
    values.font = Font(b=True, color='ffffff')
    translationIds.fill = PatternFill("solid", fgColor='000000')
    values.fill = PatternFill("solid", fgColor='000000')
    for k in j:
        value = j[k]
        ws.append([k, value])
    wb.save(outputPath)


def flatten_json(y):
    out = {}

    def flatten(x, name=''):

        # If the Nested key-value
        # pair is of dict type
        if type(x) is dict:

            for a in x:
                flatten(x[a], name + a + '.')

        # If the Nested key-value
        # pair is of list type
        elif type(x) is list:

            i = 0

            for a in x:
                flatten(a, name + str(i) + '.')
                i += 1
        else:
            out[name[:-1]] = x

    flatten(y)
    return out


if exists(jsonPath):
    print(f'[READING] {jsonPath}')
    with open(jsonPath, 'r') as f:
        data = load(f)
        normalized = flatten_json(data)
        createWorkbook(normalized)
else:
    print(f'[ERROR] no file at {jsonPath}')
