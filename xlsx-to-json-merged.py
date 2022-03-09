from datetime import datetime
from json import dumps, load
from os.path import exists

from common import checkFolders, language
from openpyxl import load_workbook
from math import floor
from pydash import merge


checkFolders()

jsonPath = './inputs/translation.json'
xlsxPath = './inputs/translation.xlsx'
outputPath = f'./outputs/merged-translations/translation-merged-{floor(datetime.timestamp(datetime.now()))}.json'

if exists(xlsxPath) and exists(jsonPath):
    print(f'[READING] {xlsxPath}')
    wb = load_workbook(xlsxPath)
    ws = wb.worksheets[0]
    translationIdsColumn = ws['A']
    valuesColumn = ws['B']
    translations = []
    for x in range(len(translationIdsColumn)):
        if (x > 0):
            translationId = translationIdsColumn[x].value
            valueTrans = valuesColumn[x].value
            translations.append(
                {'translationId': translationId, 'value': valueTrans})

    translationJson = {f'{language()}': {'translations': {}}}
    for trans in translations:
        arrayFromTranslation = trans['translationId'].split('.')
        counter = translationJson[f'{language()}']['translations']
        for i, key in enumerate(arrayFromTranslation):
            if i < len(arrayFromTranslation) - 1:
                if (counter.get(key, None) != None):
                    counter[key] = {**counter[key]}
                else:
                    counter[key] = {}
            else:
                counter[key] = trans['value']
            counter = counter[key]

    with open(jsonPath, 'r') as f:
        data = load(f)
        with open(outputPath, 'w') as outfile:
            outfile.write(dumps(merge(data, translationJson)))
else:
    print(f'[ERROR] no file at {xlsxPath}')
